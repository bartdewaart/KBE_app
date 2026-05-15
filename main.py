import os

from parapy.gui import display
from src.propulsion_system import PropulsionSystem


def load_inputs(file_path):
    """
    Integration Rule: reads mission specifications from an XLSX file.
    Returns a dict keyed by the Parameter column (row 1 = header, skipped).
    Expected columns: Parameter, Value, Unit, Description
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(
            f"Mission input file not found at '{file_path}'. "
            f"Expected location: data/input/mission.xlsx"
        )
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    specs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        param, value, *_ = row
        if param is None:
            continue
        try:
            specs[param] = float(value)
        except (TypeError, ValueError):
            specs[param] = value
    return specs


if __name__ == '__main__':

    print("========================================")
    print("  UAV Propulsion System Design Tool")
    print("========================================")
    print()

    # Step 1: load mission specifications from XLSX
    mission_file = "./data/input/mission.xlsx"
    print(f"Loading mission from: {mission_file}")
    inputs = load_inputs(mission_file)

    print()
    print("Mission parameters:")
    for name, value in inputs.items():
        print(f"  {name:<25}: {value}")
    print()
    print("Running optimization -- this may take 30-60 seconds...")
    print("========================================")
    print()

    # Step 2: instantiate PropulsionSystem with each spec as a
    # first-class Input slot so the user can edit them in the GUI tree.
    app = PropulsionSystem(
        payload_mass           = float(inputs["payload_mass"]),
        n_rotors               = int(inputs["n_rotors"]),
        safety_margin          = float(inputs["safety_margin"]),
        max_diameter           = float(inputs["max_diameter"]),
        propeller_material     = str(inputs["propeller_material"]),
        battery_energy_density = float(inputs["battery_energy_density"]),
        battery_efficiency     = float(inputs["battery_efficiency"]),
        min_endurance_min      = float(inputs["min_endurance_min"]),
        w_power                = float(inputs.get("w_power",    0.5)),
        w_mass                 = float(inputs.get("w_mass",     0.0)),
        w_endurance            = float(inputs.get("w_endurance", 0.5)),
    )

    # Step 3: initial optimization. Inside the GUI the user can edit
    # any Input and right-click the PropulsionSystem node → "Re-run
    # optimization" to refresh the optimal design.
    app.run_optimization()
    app.generate_report()

    # Step 4: launch ParaPy GUI
    display(app)